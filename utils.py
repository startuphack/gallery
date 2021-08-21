import gzip
import io
import pickle
from collections import defaultdict
from copy import copy
from datetime import datetime, timedelta

import pandas as pd
import pytz
from openpyxl import load_workbook

DEFAULT_TZ = pytz.timezone('Asia/Novosibirsk')


def parse_inventory(inventory_file, screen_file, tz=DEFAULT_TZ):
    player_ids_dict = pd.read_csv(screen_file, delimiter=';', index_col='PlayerNumber').to_dict()['PlayerId']
    inventory_df = pd.read_excel(
        inventory_file,
        parse_dates=['Дата'],
        date_parser=lambda x: tz.localize(datetime.strptime(x, '%Y-%m-%d')),
    )
    result_schedule = defaultdict(dict)
    for date, screen_name, *hours in inventory_df[['Дата', 'ID экрана'] + list(range(24))].values:
        screen_id = player_ids_dict[screen_name]
        for hour, hour_remains in enumerate(hours):
            hour_date = date + timedelta(hours=hour)
            result_schedule[screen_id][hour_date] = hour_remains

    return dict(result_schedule)


def pickle_dump(object, filename, gzip_file=True):
    """
    Сохранить объект в файл
    :param object: сохраняемый объект
    :param filename: имя файла
    :param gzip_file: сжимать ли сериализацию объекта с gzip
    """
    o_method = gzip.open if gzip_file else open

    with io.BufferedWriter(o_method(filename, 'w')) as output:
        pickle.dump(object, output, protocol=0)


def pickle_load(filename, gzip_file=True):
    """
    Загрузить объект из файла filename
    :param filename: имя файла
    :param gzip_file: является ли файл сжатым с gzip
    :return: объект, загруженный из файла
    """
    i_method = gzip.open if gzip_file else open
    with i_method(filename) as input:
        return pickle.load(input)


HOLIDAYS = [
    # 2020
    '2020-01-01', '2020-01-02', '2020-01-03', '2020-01-04', '2020-01-05', '2020-01-06', '2020-01-07', '2020-01-08',
    '2020-02-22', '2020-02-23', '2020-02-24',
    '2020-03-07', '2020-03-08', '2020-03-09',
    '2020-05-01', '2020-05-02', '2020-05-03', '2020-05-04', '2020-05-05',
    '2020-05-09', '2020-05-10', '2020-05-11',
    '2020-06-12', '2020-06-13', '2020-06-14',
    '2020-07-01',
    '2020-11-04',
    # 2021
    '2021-01-01', '2021-01-02', '2021-01-03', '2021-01-04', '2021-01-05',
    '2021-01-06', '2021-01-07', '2021-01-08', '2021-01-09', '2021-01-10',
    '2021-02-21', '2021-02-22', '2021-02-23',
    '2021-03-06', '2021-03-07', '2021-03-08',
    '2021-05-01', '2021-05-02', '2021-05-03',
    '2021-05-08', '2021-05-09', '2021-05-10',
    '2021-06-12', '2021-06-13', '2021-06-14',
    '2021-11-04', '2021-11-05', '2021-11-06', '2021-11-07',
    '2021-12-31',
]

WEEKDAYS = [
    'пн',
    'вт',
    'ср',
    'чт',
    'пт',
    'сб',
    'вс',
]


class SchedulePrinter:
    def __init__(
        self,
        template_path,
        player_details_path,
        plan_date_start,
        plan_date_stop,
        tz=DEFAULT_TZ,

    ):
        self.template_path = template_path
        self.timezone = tz
        self.player_details_path = player_details_path
        player_data = pd.read_csv(player_details_path, delimiter=';')
        self.player_name_to_ids = player_data.set_index('PlayerNumber').to_dict()['PlayerId']
        self.player_id_to_name = player_data.set_index('PlayerId').to_dict()['PlayerNumber']
        self.plan_date_start = plan_date_start
        self.plan_date_stop = plan_date_stop

        self.days = list()
        current_date = plan_date_start
        while True:
            self.days.append(current_date.date())
            current_date += timedelta(days=1)

            if current_date >= plan_date_stop:
                break

    def truncate_schedule(self, schedule):
        result = defaultdict(lambda: defaultdict(lambda: {'slots': 0, 'ots': 0}))
        for screen_id, screen_ts_items in schedule.items():
            screen_name = self.player_id_to_name[screen_id]
            for screen_ts, screen_ts_data in screen_ts_items.items():
                screen_date = datetime.fromtimestamp(screen_ts, tz=self.timezone).date()

                result[screen_name][screen_date]['slots'] += screen_ts_data['slots']
                result[screen_name][screen_date]['ots'] += screen_ts_data['ots']

        return dict(result)

    def write_schedule(self, schedule, filename=None):
        workbook = load_workbook(filename=self.template_path)
        ots_sheet = workbook['Медиаплан по OTS']

        normalized_schedule = self.truncate_schedule(schedule['schedule'])

        schedule_rows = list(normalized_schedule.items())
        background_cell = ots_sheet['b7']
        for rownum, row in enumerate(
            ots_sheet.iter_rows(min_row=6, max_row=len(schedule) + 6 - 1, min_col=2, max_col=len(self.days) + 2)):
            if rownum == 0:
                for cell, cell_date in zip(row[1:], self.days):
                    cell.value = cell_date.strftime(f'%d.%m.%Y')
                    cell.fill = copy(background_cell.fill)
            elif rownum == 1:
                for cell, cell_date in zip(row[1:], self.days):
                    cell.value = WEEKDAYS[cell_date.weekday()]
                    cell.fill = copy(background_cell.fill)
            else:
                screen_data = schedule_rows[rownum - 2]
                screen_id = screen_data[0]
                row[0].value = screen_id

                for cell, cell_date in zip(row[1:], self.days):
                    #             print(cell_date)
                    ots = screen_data[1].get(cell_date, {}).get('ots')
                    if ots:
                        cell.value = round(ots)

        slots_sheet = workbook['Медиаплан по показам']

        for rownum, row in enumerate(slots_sheet.iter_rows(min_row=3, min_col=1, max_col=50)):
            row_date = self.timezone.localize(datetime.strptime(row[0].value, '%Y-%m-%d'))

            board_name = row[3].value
            board_id = int(self.player_name_to_ids[board_name])

            board_date_data = schedule['schedule'].get(board_id)
            if board_date_data:
                total_hour_value = 0
                for hour in range(24):
                    column_stamp = int((row_date + timedelta(hours=hour)).timestamp())
                    hour_data = board_date_data.get(column_stamp)
                    if hour_data:
                        row[5 + 1 + hour].value = hour_data['slots']
                        total_hour_value += hour_data['slots']
                if total_hour_value:
                    row[5].value = total_hour_value
        if filename:
            workbook.save(filename=filename)

        return workbook
